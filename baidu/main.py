#!/usr/local/bin/python3
# coding=UTF-8

import openpyxl
import re
import datetime
from selenium import webdriver
from selenium.webdriver import Keys

g_test_batch = 1000
file_name = 'data/20210910.xlsx'
sheet_name = 'Universe2020'
title_row = 2

wb = openpyxl.load_workbook(filename=file_name, read_only=True, keep_links=False, data_only=True)
ws = wb[sheet_name]


# 计算dimensions
def calCellsRange(dim):
    [left_top, right_bottom] = dim.split(':')
    left = re.findall(r'[A-Z]+', left_top)[0]
    top = int(re.findall(r'\d+', left_top)[0])
    right = re.findall(r'[A-Z]+', right_bottom)[0]
    bottom = int(re.findall(r'\d+', right_bottom)[0])
    return {'left': left, 'top': top, 'right': right, 'bottom': bottom}


dim = calCellsRange(ws.calculate_dimension())


# 总行数
def calDataRowsCount(dim):
    return dim['bottom'] - (title_row + 1) + 1


# data_rows_count = 0
if g_test_batch > 0:
    data_rows_count = g_test_batch
else:
    data_rows_count = calDataRowsCount(dim)


driver = webdriver.Chrome(executable_path='lib/chromedriver')
driver.implicitly_wait(10)     # seconds
driver.get('http://www.baidu.com')


def searchOneName(id, name, driver):
    element = driver.find_element_by_id('kw')
    element.send_keys(name)
    element.send_keys(Keys.ENTER)

    is_full_match = 0
    elems = driver.find_elements_by_xpath('//div[contains(@class,"result-op") or contains(@class,"result")]/h3/a')
    for ele in elems:
        if name in ele.text:
            is_full_match = 1
            break

    driver.back()
    return (id, name, is_full_match, 'Others')


target_list = []
rows = ws.iter_rows(title_row + 1, title_row + data_rows_count)
for row in rows:
    print('alfred')
    t_pha_id = row[5].value
    t_pha_name = row[7].value
    target_list.append(searchOneName(t_pha_id, t_pha_name, driver))

driver.quit()

def save2Result(target_list, result_file):
    # 创建一个workbook对象，而且会在workbook中至少创建一个表worksheet
    wb = openpyxl.Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    ws = wb.active
    title_data = ('PHAID', 'HOSP_NAME', 'FULL_MATCH', 'OTHERS')
    target_list.insert(0, title_data)
    rows = len(target_list)
    cols = len(target_list[0])
    for i in range(rows):
        for j in range(cols):
            ws.cell(row=i + 1, column=j + 1).value = target_list[i][j]
    wb.save(filename=result_file)


file_name = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
save2Result(target_list, file_name + '.xlsx')
