#!/usr/local/bin/python3
# coding=UTF-8

import openpyxl
import re
import requests
import time
import datetime
import random
import urllib
import ssl
from time import perf_counter
from urllib.parse import quote
import string
from bs4 import BeautifulSoup

g_test_batch = 10
file_name = 'data/20210910.xlsx'
sheet_name = 'Universe2020'
title_row = 2
g_count = 0

wb = openpyxl.load_workbook(filename=file_name, read_only=True, keep_links=False, data_only=True)
ws = wb[sheet_name]


# get random ip
def queryRandomIpList():
    url = 'http://www.xiladaili.com/'
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Safari/537.36'}
    html = requests.get(url=url, headers=headers).text
    soup = BeautifulSoup(html, 'html.parser')
    ips = soup.find_all('tr')
    ip_list = []
    for ip_info in ips:
        tds = ip_info.find_all('td')
        if len(tds) == 8:
            ip_list.append('http://' + tds[0].text)
    print("代理列表抓取成功.")

    return ip_list


# 计算dimensions
def calCellsRange(dim):
    [left_top, right_bottom] = dim.split(':')
    left = re.findall(r'[A-Z]+', left_top)[0]
    top = int(re.findall(r'\d+', left_top)[0])
    right = re.findall(r'[A-Z]+', right_bottom)[0]
    bottom = int(re.findall(r'\d+', right_bottom)[0])
    return {'left': left, 'top': top, 'right': right, 'bottom': bottom}


dim = calCellsRange(ws.calculate_dimension())


# 总行数
def calDataRowsCount(dim):
    return dim['bottom'] - (title_row + 1) + 1


# data_rows_count = 0
if g_test_batch > 0:
    data_rows_count = g_test_batch
else:
    data_rows_count = calDataRowsCount(dim)

ip_list = queryRandomIpList()
print(ip_list)
user_agent = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:61.0) Gecko/20100101 Firefox/61.0",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)",
    "Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)",
    "Mozilla/4.0 (compatible; MSIE 7.0b; Windows NT 5.2; .NET CLR 1.1.4322; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 3.0.04506.30)",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN) AppleWebKit/523.15 (KHTML, like Gecko, Safari/419.3) Arora/0.3 (Change: 287 c9dfb30)",
    "Mozilla/5.0 (X11; U; Linux; en-US) AppleWebKit/527+ (KHTML, like Gecko, Safari/419.3) Arora/0.6",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.2pre) Gecko/20070215 K-Ninja/2.1.1",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9) Gecko/20080705 Firefox/3.0 Kapiko/3.0",
    "Mozilla/5.0 (X11; Linux i686; U;) Gecko/20070322 Kazehakase/0.4.5"]
sk = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ123456789'
sp = '123456789'
fm = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'


target_list = []
rows = ws.iter_rows(title_row + 1, title_row + data_rows_count)
for row in rows:
    t_pha_id = row[5].value
    t_pha_name = row[7].value

    # url = 'http://www.baidu.com/s?wd=' + t_pha_name + '&usm=3&rsv_idx=2&rsv_page=1'
    url = 'https://cn.bing.com/search?q=' + t_pha_name + '&sk=' + \
          ''.join(random.choices(sk, k=3)) + '&sc=8-7' + '&cvid=' + \
          ''.join(random.choices(sk, k=32)) + '&FORM=' + \
          ''.join(random.choices(fm, k=4)) + '&sp=' + \
          random.choice(sp)
    url = quote(url, safe=string.printable)
    print(url)
    headers = {'User-Agent': random.choice(user_agent), "Upgrade-Insecure-Requests": "1",
               "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
               "Accept-Encoding": "gzip, deflate, sdch",
               "Accept-Language": "zh-CN,zh;q=0.8,en;q=0.6",
               "Cache-Control": "no-cache"}

    begin = perf_counter()
    httpproxy_handler = urllib.request.ProxyHandler({'http': random.choice(ip_list)})
    opener = urllib.request.build_opener(httpproxy_handler)
    request = urllib.request.Request(url)
    urllib.request.install_opener(opener)
    # response = opener.open(request)
    response = urllib.request.urlopen(request, context=ssl._create_unverified_context())
    # print(response.read().decode('utf-8'))
    html = response.read().decode('utf-8')
    # print(html)
    end = perf_counter()
    print("scraper {0:.2f}s".format(end - begin))
    # html = requests.get(url, proxies={'http': random.choice(ip_list)}, headers=headers).text
    # print(html)

    tmp = ()
    soup = BeautifulSoup(html, 'html.parser')
    # 缺反爬虫处理
    bt = soup.find_all('div', class_='b_title')
    if len(bt) == 0:
        print('error')
        time.sleep(10)

    for div in soup.find_all('div', class_='b_title'):
        hits = div.find_all('h2')
        for hit in hits:
            if t_pha_name in hit.text:
                tmp = (t_pha_id, t_pha_name, 1, 'other')
                break
            else:
                tmp = (t_pha_id, t_pha_name, 0, 'other')

    target_list.append(tmp)
    # time.sleep(10)


def save2Result(target_list, result_file):
    # 创建一个workbook对象，而且会在workbook中至少创建一个表worksheet
    wb = openpyxl.Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    ws = wb.active
    title_data = ('PHAID', 'HOSP_NAME', 'FULL_MATCH', 'OTHERS')
    target_list.insert(0, title_data)
    rows = len(target_list)
    cols = len(target_list[0])
    for i in range(rows):
        for j in range(cols):
            ws.cell(row=i + 1, column=j + 1).value = target_list[i][j]
    wb.save(filename=result_file)


file_name = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
save2Result(target_list, file_name)
